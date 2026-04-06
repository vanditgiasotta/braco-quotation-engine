"""
Braco Output Generator — Cluster 3
=====================================
Generates a client-ready Excel quotation that exactly matches
the Sample 5 / Braco standard format.

RELEASE GATE: Output is blocked if validation release_allowed = False.
No BLOCKED lines are ever written to the output file.
WARNING lines are written with a visible flag note.

Connects to:
  braco_engine.py    → SelectionResult, calculate_prices
  braco_validator.py → ValidationReport, QuoteValidationSummary
  braco_parser.py    → ParsedLine, parsed_line_to_line_item (optional, for full pipeline)

Usage (standalone test):
  python3 braco_output.py

Usage (full pipeline):
  from braco_output import generate_quotation, QuoteConfig
"""

import sys, os, re
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from braco_engine import LineItem, SelectionResult, run_selection, calculate_prices
from braco_validator import validate, ValidationReport, QuoteValidationSummary, validate_quote

# ══════════════════════════════════════════════════════════════════════
# SECTION 1 — QUOTE CONFIGURATION
# Everything the employee fills in before generating output.
# ══════════════════════════════════════════════════════════════════════

@dataclass
class QuoteConfig:
    quote_ref: str          # e.g. "QT0000082"
    quote_date: str         # e.g. "06.04.2026"
    client_name: str
    client_address: str     # optional, shown on quote
    project_name: str       # e.g. "KXT-26011 Tunisia"
    section_label: str      # e.g. "16.2" — matches client BOQ numbering
    section_title: str      # e.g. "LT Cable Termination"
    section_description: str  # the spec paragraph under section header
    discount_pct: float     # e.g. 46.0
    generated_by: str       # employee name
    approved_by: str        # director / approver
    currency: str = "INR"   # INR or USD or EUR
    is_export: bool = False
    include_validation_sheet: bool = True
    # Terms flags (driven by what's in the quote)
    has_glands: bool = True
    has_lugs: bool = True


# ══════════════════════════════════════════════════════════════════════
# SECTION 2 — STYLE PALETTE
# Exactly matches Sample 5 visual style.
# ══════════════════════════════════════════════════════════════════════

# Colours (hex, no #)
COL_HEADER_DARK  = "1F3864"   # dark navy — main section headers
COL_HEADER_MID   = "2E75B6"   # medium blue — column headers
COL_HEADER_LIGHT = "BDD7EE"   # light blue — sub-headers
COL_ALT_ROW      = "EBF3FB"   # very light blue — alternating rows
COL_WHITE        = "FFFFFF"
COL_TOTAL_BG     = "FFF2CC"   # amber — totals row
COL_WARN_BG      = "FFE6CC"   # orange — warning rows
COL_BLOCKED_BG   = "FFD7D7"   # light red — blocked rows (shouldn't appear)
COL_NOTE_BG      = "F2F2F2"   # light grey — notes / T&C
COL_BLACK        = "000000"
COL_WHITE_TEXT   = "FFFFFF"
COL_DARK_TEXT    = "1F1F1F"
COL_BLUE_TEXT    = "1F3864"

FONT_MAIN  = "Arial"
FONT_SIZE  = 9
FONT_HDR   = 10


def _font(bold=False, size=FONT_SIZE, color=COL_DARK_TEXT, name=FONT_MAIN, italic=False):
    return Font(bold=bold, size=size, color=color, name=name, italic=italic)


def _fill(color):
    return PatternFill("solid", start_color=color, fgColor=color)


def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(style="thin", color="AAAAAA"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _thin():
    return _border("thin", "CCCCCC")


def _thick():
    return _border("medium", "1F3864")


# ══════════════════════════════════════════════════════════════════════
# SECTION 3 — COLUMN MAP
# Exact column layout from Sample 5 analysis.
# All columns are 1-indexed for openpyxl.
# ══════════════════════════════════════════════════════════════════════

#  Col:  A   B   C          D      E    F     G           H        I       J           K        L      M           N        O
#  Idx:  1   2   3          4      5    6     7           8        9       10          11       12     13          14       15
#  Use:  -  Ref Description Unit  Qty  OD  Gland Desc  Cat No  List₹  Lug FC Desc  Cat No  List₹  Lug HC Desc  Cat No  List₹

COL_IDX      = 1   # A — row index letter (a, b, c...)
COL_REF      = 2   # B — section ref (16.2, 16.3...)
COL_DESC     = 3   # C — cable description
COL_UNIT     = 4   # D — unit (NO'S)
COL_QTY      = 5   # E — quantity
COL_OD       = 6   # F — cable OD (mm)
COL_G_DESC   = 7   # G — gland description
COL_G_CAT    = 8   # H — gland cat no
COL_G_PRICE  = 9   # I — gland list price
COL_LF_DESC  = 10  # J — full-core lug description
COL_LF_CAT   = 11  # K — full-core lug cat no
COL_LF_PRICE = 12  # L — full-core lug list price
COL_LH_DESC  = 13  # M — half-core lug description
COL_LH_CAT   = 14  # N — half-core lug cat no
COL_LH_PRICE = 15  # O — half-core lug list price

TOTAL_COLS = 15

# Column widths (characters)
COL_WIDTHS = {
    1: 4,   # A idx
    2: 6,   # B ref
    3: 36,  # C description — widest
    4: 6,   # D unit
    5: 7,   # E qty
    6: 7,   # F OD
    7: 38,  # G gland description
    8: 12,  # H gland cat no
    9: 10,  # I gland price
    10: 36, # J lug FC desc
    11: 12, # K lug FC cat
    12: 10, # L lug FC price
    13: 36, # M lug HC desc
    14: 12, # N lug HC cat
    15: 10, # O lug HC price
}


def _set_col_widths(ws):
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ══════════════════════════════════════════════════════════════════════
# SECTION 4 — ROW WRITERS
# Each writes one logical row type to the worksheet.
# All take `ws` and `row` (1-indexed row number) and return next_row.
# ══════════════════════════════════════════════════════════════════════

def _cell(ws, row, col, value=None, font=None, fill=None, align=None, border=None,
          number_format=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font:    c.font   = font
    if fill:    c.fill   = fill
    if align:   c.alignment = align
    if border:  c.border  = border
    if number_format: c.number_format = number_format
    return c


def write_quote_header(ws, row: int, config: QuoteConfig) -> int:
    """Row 1: Quote ref number + date (right-aligned)"""
    ws.row_dimensions[row].height = 18
    c = ws.cell(row=row, column=COL_DESC)
    c.value = f"Braco Quotation Ref NO : {config.quote_ref} / Dt.{config.quote_date}"
    c.font = _font(bold=True, size=10, color=COL_BLUE_TEXT)
    c.alignment = _align("left")
    ws.merge_cells(start_row=row, start_column=COL_DESC,
                   end_row=row,   end_column=COL_OD)
    return row + 1


def write_braco_cat_label(ws, row: int) -> int:
    """Row 2: 'BRACO CAT SELECTION' centred across gland+lug columns"""
    ws.row_dimensions[row].height = 16
    c = ws.cell(row=row, column=COL_G_DESC)
    c.value = "BRACO CAT SELECTION"
    c.font = _font(bold=True, size=10, color=COL_WHITE_TEXT)
    c.fill = _fill(COL_HEADER_DARK)
    c.alignment = _align("center")
    ws.merge_cells(start_row=row, start_column=COL_G_DESC,
                   end_row=row,   end_column=COL_LH_PRICE)
    return row + 1


def write_section_header(ws, row: int, ref: str, title: str, description: str) -> int:
    """Section header row: ref + title"""
    ws.row_dimensions[row].height = 18
    # Col B: section ref
    _cell(ws, row, COL_REF, ref,
          font=_font(bold=True, size=10, color=COL_WHITE_TEXT),
          fill=_fill(COL_HEADER_DARK),
          align=_align("center"))
    # Col C+: title
    ws.merge_cells(start_row=row, start_column=COL_DESC,
                   end_row=row,   end_column=TOTAL_COLS)
    _cell(ws, row, COL_DESC, title,
          font=_font(bold=True, size=10, color=COL_WHITE_TEXT),
          fill=_fill(COL_HEADER_DARK),
          align=_align("left"))
    row += 1

    # Description row if provided
    if description:
        ws.row_dimensions[row].height = 30
        ws.merge_cells(start_row=row, start_column=COL_IDX,
                       end_row=row,   end_column=TOTAL_COLS)
        _cell(ws, row, COL_IDX, description,
              font=_font(italic=True, size=8, color=COL_DARK_TEXT),
              fill=_fill("F8F8F8"),
              align=_align("left", wrap=True))
        row += 1
    return row


def write_col_headers(ws, row: int, has_glands: bool, has_lugs: bool) -> int:
    """Column header row: OD | Gland | Full Core Lug | Half Core Lug"""
    ws.row_dimensions[row].height = 28

    def hdr(col, text, span=1):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
        _cell(ws, row, col, text,
              font=_font(bold=True, size=9, color=COL_WHITE_TEXT),
              fill=_fill(COL_HEADER_MID),
              align=_align("center", wrap=True),
              border=_thin())

    hdr(COL_IDX,  "")
    hdr(COL_REF,  "")
    hdr(COL_DESC, "Description", 1)
    hdr(COL_UNIT, "Unit")
    hdr(COL_QTY,  "Qty")
    if has_glands:
        hdr(COL_OD,     "OD (mm)")
        hdr(COL_G_DESC, "Description",  1)
        hdr(COL_G_CAT,  "Cat No")
        hdr(COL_G_PRICE,"Price (₹)")
    if has_lugs:
        hdr(COL_LF_DESC, "Description",  1)
        hdr(COL_LF_CAT,  "Cat No")
        hdr(COL_LF_PRICE,"Price (₹)")
        hdr(COL_LH_DESC, "Description",  1)
        hdr(COL_LH_CAT,  "Cat No")
        hdr(COL_LH_PRICE,"Price (₹)")

    # Sub-headers for lug columns
    row += 1
    ws.row_dimensions[row].height = 16
    sub = _fill(COL_HEADER_LIGHT)
    sub_font = _font(bold=True, size=8, color=COL_BLUE_TEXT)
    if has_lugs:
        ws.merge_cells(start_row=row, start_column=COL_LF_DESC, end_row=row, end_column=COL_LF_PRICE)
        _cell(ws, row, COL_LF_DESC, "FULL CORE", font=sub_font, fill=sub, align=_align("center"))
        ws.merge_cells(start_row=row, start_column=COL_LH_DESC, end_row=row, end_column=COL_LH_PRICE)
        _cell(ws, row, COL_LH_DESC, "HALF CORE", font=sub_font, fill=sub, align=_align("center"))

    return row + 1


def write_data_row(ws, row: int, idx_letter: str, item: LineItem,
                   result: SelectionResult, validation: ValidationReport,
                   alt: bool) -> int:
    """One data row per cable line."""
    ws.row_dimensions[row].height = 20
    bg = COL_WARN_BG if validation.final_status == "NEEDS_REVIEW" else (
         COL_ALT_ROW  if alt else COL_WHITE)
    fill = _fill(bg)
    font = _font()
    align_c = _align("center")
    align_l = _align("left", wrap=True)
    b = _thin()

    # Col A — index letter
    _cell(ws, row, COL_IDX, idx_letter, font=font, fill=fill, align=align_c, border=b)
    # Col B — blank (section ref used in header only)
    _cell(ws, row, COL_REF, "", fill=fill, border=b)
    # Col C — cable description
    _cell(ws, row, COL_DESC, item.description, font=font, fill=fill, align=align_l, border=b)
    # Col D — unit
    _cell(ws, row, COL_UNIT, "NO'S", font=font, fill=fill, align=align_c, border=b)
    # Col E — quantity
    _cell(ws, row, COL_QTY, item.qty, font=font, fill=fill, align=align_c, border=b)

    # Gland columns
    if result.gland:
        od_val = result.od_used if result.od_used else ""
        _cell(ws, row, COL_OD, od_val, font=font, fill=fill, align=align_c, border=b)
        _cell(ws, row, COL_G_DESC, result.gland["description"],
              font=font, fill=fill, align=align_l, border=b)
        _cell(ws, row, COL_G_CAT, result.gland["cat_no"],
              font=_font(bold=True), fill=fill, align=align_c, border=b)
        _cell(ws, row, COL_G_PRICE, result.gland["list_price"],
              font=font, fill=fill, align=align_c, border=b,
              number_format='#,##0.00')
    else:
        for c in [COL_OD, COL_G_DESC, COL_G_CAT, COL_G_PRICE]:
            _cell(ws, row, c, "", fill=fill, border=b)

    # Full-core lug columns
    if result.lug_full:
        _cell(ws, row, COL_LF_DESC, result.lug_full["description"],
              font=font, fill=fill, align=align_l, border=b)
        _cell(ws, row, COL_LF_CAT, result.lug_full["cat_no"],
              font=_font(bold=True), fill=fill, align=align_c, border=b)
        _cell(ws, row, COL_LF_PRICE, result.lug_full["list_price"],
              font=font, fill=fill, align=align_c, border=b, number_format='#,##0.00')
    else:
        for c in [COL_LF_DESC, COL_LF_CAT, COL_LF_PRICE]:
            _cell(ws, row, c, "", fill=fill, border=b)

    # Half-core lug columns
    if result.lug_half:
        _cell(ws, row, COL_LH_DESC, result.lug_half["description"],
              font=font, fill=fill, align=align_l, border=b)
        _cell(ws, row, COL_LH_CAT, result.lug_half["cat_no"],
              font=_font(bold=True), fill=fill, align=align_c, border=b)
        _cell(ws, row, COL_LH_PRICE, result.lug_half["list_price"],
              font=font, fill=fill, align=align_c, border=b, number_format='#,##0.00')
    else:
        for c in [COL_LH_DESC, COL_LH_CAT, COL_LH_PRICE]:
            _cell(ws, row, c, "", fill=fill, border=b)

    # Warning flag in description cell if needed
    if validation.final_status == "NEEDS_REVIEW":
        ws.cell(row=row, column=COL_DESC).value += "  ⚠ REVIEW"

    return row + 1


def write_pricing_block(ws, row: int, data_rows: list, discount_pct: float, config: QuoteConfig) -> int:
    """
    Pricing summary block below the data rows.
    Shows: Discount %, Net Rate per item, Grand Total Net.
    Uses Excel formulas — not hardcoded Python values.

    data_rows: list of (excel_row_number, item, result, prices)
    """
    ws.row_dimensions[row].height = 4
    row += 1  # spacer

    # Discount line
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=COL_IDX, end_row=row, end_column=COL_G_DESC)
    _cell(ws, row, COL_IDX,
          f"Discount Applied:  {discount_pct:.0f}%  (Net Rates)",
          font=_font(bold=True, size=9, color=COL_BLUE_TEXT),
          fill=_fill(COL_TOTAL_BG),
          align=_align("right"))
    row += 1

    # Net price per item header
    ws.row_dimensions[row].height = 16
    _cell(ws, row, COL_G_PRICE, f"Net Price (₹)\n@{discount_pct:.0f}% disc",
          font=_font(bold=True, size=8, color=COL_WHITE_TEXT),
          fill=_fill(COL_HEADER_MID),
          align=_align("center", wrap=True))
    _cell(ws, row, COL_LF_PRICE, "FC Lug Net",
          font=_font(bold=True, size=8, color=COL_WHITE_TEXT),
          fill=_fill(COL_HEADER_MID),
          align=_align("center"))
    _cell(ws, row, COL_LH_PRICE, "HC Lug Net",
          font=_font(bold=True, size=8, color=COL_WHITE_TEXT),
          fill=_fill(COL_HEADER_MID),
          align=_align("center"))
    row += 1

    # Net price rows — Excel formula: list_price_cell * (1 - disc/100)
    mult_str = f"*(1-{discount_pct}/100)"
    net_price_cells_g  = []
    net_price_cells_lf = []
    net_price_cells_lh = []

    for (data_row, item, result, _) in data_rows:
        ws.row_dimensions[row].height = 18

        # Gland net
        if result.gland:
            g_price_cell = f"{get_column_letter(COL_G_PRICE)}{data_row}"
            qty_cell     = f"{get_column_letter(COL_QTY)}{data_row}"
            g_net_col = get_column_letter(COL_G_PRICE)
            c = ws.cell(row=row, column=COL_G_PRICE)
            c.value = f"={g_price_cell}{mult_str}*{qty_cell}"
            c.font = _font()
            c.alignment = _align("center")
            c.number_format = '#,##0.00'
            c.border = _thin()
            net_price_cells_g.append(f"{get_column_letter(COL_G_PRICE)}{row}")

        # FC lug net
        if result.lug_full:
            lf_price_cell = f"{get_column_letter(COL_LF_PRICE)}{data_row}"
            qty_cell      = f"{get_column_letter(COL_QTY)}{data_row}"
            c = ws.cell(row=row, column=COL_LF_PRICE)
            c.value = f"={lf_price_cell}{mult_str}*{qty_cell}"
            c.font = _font()
            c.alignment = _align("center")
            c.number_format = '#,##0.00'
            c.border = _thin()
            net_price_cells_lf.append(f"{get_column_letter(COL_LF_PRICE)}{row}")

        # HC lug net
        if result.lug_half:
            lh_price_cell = f"{get_column_letter(COL_LH_PRICE)}{data_row}"
            qty_cell      = f"{get_column_letter(COL_QTY)}{data_row}"
            c = ws.cell(row=row, column=COL_LH_PRICE)
            c.value = f"={lh_price_cell}{mult_str}*{qty_cell}"
            c.font = _font()
            c.alignment = _align("center")
            c.number_format = '#,##0.00'
            c.border = _thin()
            net_price_cells_lh.append(f"{get_column_letter(COL_LH_PRICE)}{row}")

        row += 1

    # Grand total row
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=COL_IDX, end_row=row, end_column=COL_G_DESC)
    _cell(ws, row, COL_IDX, f"GRAND TOTAL NET  (@ {discount_pct:.0f}% Discount)",
          font=_font(bold=True, size=10, color=COL_DARK_TEXT),
          fill=_fill(COL_TOTAL_BG),
          align=_align("right"),
          border=_thick())

    # Sum formula for each price column
    for col, cells in [(COL_G_PRICE, net_price_cells_g),
                       (COL_LF_PRICE, net_price_cells_lf),
                       (COL_LH_PRICE, net_price_cells_lh)]:
        if cells:
            formula = "=" + "+".join(cells)
            c = ws.cell(row=row, column=col)
            c.value = formula
            c.font = _font(bold=True, size=10)
            c.fill = _fill(COL_TOTAL_BG)
            c.alignment = _align("center")
            c.number_format = '#,##0.00'
            c.border = _thick()

    row += 2
    return row


def write_notes_and_tc(ws, row: int, config: QuoteConfig) -> int:
    """Terms & conditions block — conditional on what's in the quote."""
    note_font = _font(size=8, color="444444")
    note_fill = _fill(COL_NOTE_BG)
    note_align = _align("left", wrap=True)

    def note_row(text, bold=False, height=14):
        ws.row_dimensions[row].height = height
        ws.merge_cells(start_row=row, start_column=COL_IDX, end_row=row, end_column=TOTAL_COLS)
        c = ws.cell(row=row, column=COL_IDX)
        c.value = text
        c.font = _font(bold=bold, size=8, color="333333")
        c.fill = note_fill
        c.alignment = note_align
        return row + 1

    # OD note — only if any OD was from Polycab reference (not client-stated)
    r = row
    r = note_row("")
    if config.has_glands:
        r = note_row(
            "NOTE: Gland selection is done as per Thread size / Cable OD data available with us. "
            "Please confirm actual Cable OD before placing order as Gland size may change as per cable OD.",
            bold=False, height=30,
        )

    r = note_row("")
    r = note_row("Terms & Conditions :", bold=True, height=16)

    tc_lines = [
        "The above rates are NET rates & valid only for full quantity order.",
        "We have attached our Technical Sheets, kindly approve the same before placement of any order. "
        "Later on no replacement will be entertained.",
        "The prices are EX-BRACO Works, Freight on actuals to be borne by you.",
    ]
    if config.has_glands:
        tc_lines.append("The above Glands are quoted without PVC Shroud & Earth tag.")
    tc_lines += [
        "GST @ 18% shall be charged extra.",
        "GST Registration No.: 27AAECB0129N1ZO",
        "HSN Code: Cable Glands: 85389000 ; GST @ 18%,  Cable Lugs: 85369090 ; GST @ 18%",
        "Any Attestation of documents, Type test & inspection by outside agencies — charges extra.",
        "Internal Test Certificates if required, shall be sent along with Original Invoice.",
        "Dispatch: Whatever in Stock can be dispatched in 4 to 6 working days & balance within 3–4 weeks.",
        "Payment: 20% Advance Payment against PO and balance against Proforma Invoice before dispatch.",
        "This offer is valid for 7 days.",
    ]
    if config.is_export:
        tc_lines.append(
            "Export pricing: Net rates as mutually agreed. Prices subject to exchange rate fluctuations."
        )

    for i, line in enumerate(tc_lines, 1):
        r = note_row(f"{i}.  {line}", height=22)

    r += 1
    # Generated / approved by
    ws.row_dimensions[r].height = 16
    ws.merge_cells(start_row=r, start_column=COL_IDX, end_row=r, end_column=TOTAL_COLS)
    c = ws.cell(row=r, column=COL_IDX)
    c.value = (f"Generated by: {config.generated_by}    |    "
               f"Approved by: {config.approved_by}    |    "
               f"Date: {config.quote_date}")
    c.font = _font(italic=True, size=8, color="666666")
    c.fill = _fill(COL_WHITE)
    c.alignment = _align("right")
    return r + 1


# ══════════════════════════════════════════════════════════════════════
# SECTION 5 — VALIDATION SUMMARY SHEET
# Second sheet: human-readable audit trail for internal use.
# ══════════════════════════════════════════════════════════════════════

def write_validation_sheet(wb, summary: QuoteValidationSummary, config: QuoteConfig):
    ws = wb.create_sheet("Validation Audit")
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 50

    row = 1

    def hdr_cell(r, c, v, bg=COL_HEADER_DARK):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = _font(bold=True, color=COL_WHITE_TEXT)
        cell.fill = _fill(bg)
        cell.alignment = _align("center")
        cell.border = _thin()
        return cell

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1, value=f"VALIDATION AUDIT — {config.quote_ref}")
    c.font = _font(bold=True, size=12, color=COL_WHITE_TEXT)
    c.fill = _fill(COL_HEADER_DARK)
    c.alignment = _align("center")
    row = 2

    # Summary bar
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    verdict = summary.release_verdict
    verdict_bg = "C6EFCE" if summary.release_allowed else "FFC7CE"
    c = ws.cell(row=row, column=1, value=verdict)
    c.font = _font(bold=True, size=9)
    c.fill = _fill(verdict_bg)
    c.alignment = _align("left")
    row += 1

    # Stats
    for label, val in [
        ("Total Lines", summary.total_lines),
        ("Approved", summary.approved),
        ("Needs Review", summary.needs_review),
        ("Blocked", summary.blocked),
        ("Grand Total Net", f"₹{summary.grand_total_net:,.2f}"),
    ]:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, size=8)
        ws.cell(row=row, column=2, value=val).font = _font(size=8)
        row += 1

    row += 1
    # Column headers for detail
    for c_idx, label in [(1,"Line"),(2,"Description"),(3,"Status"),(4,"Confidence"),(5,"Check ID"),(6,"Issue")]:
        hdr_cell(row, c_idx, label, COL_HEADER_MID)
    row += 1

    # Per-line detail
    status_bg = {"APPROVED":"C6EFCE","NEEDS_REVIEW":"FFEB9C","BLOCKED":"FFC7CE"}
    for report in summary.reports:
        bg = status_bg.get(report.final_status, COL_WHITE)
        ws.cell(row=row, column=1, value=report.line_no).font = _font(size=8)
        ws.cell(row=row, column=2, value=report.description[:40]).font = _font(size=8)
        c = ws.cell(row=row, column=3, value=report.final_status)
        c.font = _font(bold=True, size=8)
        c.fill = _fill(bg)
        c.alignment = _align("center")
        ws.cell(row=row, column=4, value=f"{report.confidence_score:.0%}").font = _font(size=8)

        # Inline all checks that fired
        issues = [f"[{ch['severity']}] {ch['title']}" for ch in report.checks
                  if ch['severity'] not in ('PASS',)]
        for i, issue in enumerate(issues):
            if i == 0:
                ws.cell(row=row, column=5, value=report.checks[i].get('check_id',''))
                ws.cell(row=row, column=6, value=issue).font = _font(size=8)
            else:
                row += 1
                ws.cell(row=row, column=5, value='').font = _font(size=8)
                ws.cell(row=row, column=6, value=issue).font = _font(size=8)
        row += 1


# ══════════════════════════════════════════════════════════════════════
# SECTION 6 — MAIN ORCHESTRATOR
# generate_quotation() is the single function called by the pipeline.
# ══════════════════════════════════════════════════════════════════════

def generate_quotation(
    items: list,               # list[LineItem] — from parser
    results: list,             # list[SelectionResult] — from engine
    validation_summary: QuoteValidationSummary,
    config: QuoteConfig,
    output_path: str,
) -> dict:
    """
    Generate the Excel quotation file.
    Returns: {"ok": bool, "path": str, "reason": str}

    RELEASE GATE: If validation_summary.release_allowed is False,
    no file is created and the function returns immediately.
    """

    # ── RELEASE GATE ─────────────────────────────────────────────────
    if not validation_summary.release_allowed:
        return {
            "ok": False,
            "path": None,
            "reason": (
                f"Output BLOCKED: {validation_summary.blocked} line(s) have critical validation errors. "
                "Resolve all BLOCK items before generating output. "
                f"Verdict: {validation_summary.release_verdict}"
            ),
        }

    # ── Only include non-BLOCKED lines ───────────────────────────────
    valid_reports = {r.line_no: r for r in validation_summary.reports}
    approved_pairs = [
        (item, result) for item, result in zip(items, results)
        if valid_reports.get(item.line_no, None) and
           valid_reports[item.line_no].final_status != "BLOCKED"
    ]

    # ── Build workbook ────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws)

    row = 1
    row = write_quote_header(ws, row, config)
    row = write_braco_cat_label(ws, row)
    row += 1  # spacer

    # Group by section
    from collections import defaultdict
    sections = defaultdict(list)
    for item, result in approved_pairs:
        sections[item.section].append((item, result))

    data_rows_for_pricing = []   # (excel_row, item, result, prices) — all sections
    has_glands = any(r.gland for _, r in approved_pairs)
    has_lugs   = any(r.lug_full or r.lug_half for _, r in approved_pairs)
    config.has_glands = has_glands
    config.has_lugs   = has_lugs

    idx_counters = {}  # section → letter counter
    alpha = "abcdefghijklmnopqrstuvwxyz"

    for sec_name, pairs in sections.items():
        # Section header
        row = write_section_header(ws, row, config.section_label, sec_name,
                                   config.section_description if sec_name == list(sections.keys())[0]
                                   else "")
        row = write_col_headers(ws, row, has_glands, has_lugs)

        idx_counters[sec_name] = 0

        for item, result in pairs:
            idx_letter = alpha[idx_counters[sec_name] % 26]
            idx_counters[sec_name] += 1
            report = valid_reports.get(item.line_no)
            prices = calculate_prices(result, config.discount_pct)
            data_row = row
            row = write_data_row(ws, row, idx_letter, item, result, report,
                                 alt=(idx_counters[sec_name] % 2 == 0))
            data_rows_for_pricing.append((data_row, item, result, prices))

        row += 1  # gap between sections

    # Pricing block
    row = write_pricing_block(ws, row, data_rows_for_pricing, config.discount_pct, config)

    # Notes and T&C
    row = write_notes_and_tc(ws, row, config)

    # Freeze panes at row 4 (below headers)
    ws.freeze_panes = "C4"

    # Validation audit sheet
    if config.include_validation_sheet:
        write_validation_sheet(wb, validation_summary, config)

    # Save
    wb.save(output_path)
    return {"ok": True, "path": output_path, "reason": "OK"}


# ══════════════════════════════════════════════════════════════════════
# SECTION 7 — FULL PIPELINE TEST
# Runs: parse → engine → validate → output
# Uses real Tunisia BOQ data.
# ══════════════════════════════════════════════════════════════════════

def run_output_test():
    print("\n" + "="*68)
    print("  BRACO OUTPUT GENERATOR — CLUSTER 3 — FULL PIPELINE TEST")
    print("="*68 + "\n")

    # ── Step 1: Define test line items (Tunisia BOQ aggregated) ───────
    # Using a representative subset matching Sample 5 structure
    test_items = [
        # LV Power Cable Terminations (glands + lugs)
        LineItem(1, "3CX35 SQ.MM",   3,   35,  26, od_stated=25.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(2, "3.5CX25 SQ.MM", 3.5, 25,   8, od_stated=23.5,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(3, "3.5CX35 SQ.MM", 3.5, 35,  42, od_stated=26.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(4, "3.5CX50 SQ.MM", 3.5, 50,  12, od_stated=30.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(5, "3.5CX95 SQ.MM", 3.5, 95,  90, od_stated=36.5,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(6, "3.5CX120 SQ.MM",3.5,120,  54, od_stated=40.5,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(7, "3.5CX185 SQ.MM",3.5,185,   8, od_stated=50.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(8, "3.5CX240 SQ.MM",3.5,240,   4, od_stated=55.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(9, "3.5CX300 SQ.MM",3.5,300,  52, od_stated=61.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(10,"4CX16 SQ.MM",   4,   16,  10, od_stated=23.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(11,"4CX10 SQMM",    4,   10,  10, od_stated=20.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(12,"4CX25 SQ.MM",   4,   25,  10, od_stated=24.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        LineItem(13,"4CX35 SQ.MM",   4,   35,  10, od_stated=27.0,  gland_pref="BPW", needs_gland=True, needs_lug=True,  section="LT Cable Termination"),
        # Control Cable Glands only
        LineItem(14,"4Cx10 Sq.mm CU PVC", 4,10,524, gland_pref="BPW", needs_gland=True, needs_lug=False, section="Control Cable Glands"),
        LineItem(15,"4Cx6 Sq.mm CU PVC",  4, 6,1008, gland_pref="BPW", needs_gland=True, needs_lug=False, section="Control Cable Glands"),
    ]

    DISCOUNT = 46.0

    # ── Step 2: Run selection engine ──────────────────────────────────
    print("  [1/4] Running selection engine...")
    results = [run_selection(item) for item in test_items]

    # ── Step 3: Run validation ────────────────────────────────────────
    print("  [2/4] Running validation...")
    val_summary = validate_quote(test_items, results, DISCOUNT)
    print(f"        {val_summary.approved} approved, "
          f"{val_summary.needs_review} needs review, "
          f"{val_summary.blocked} blocked")
    print(f"        Release allowed: {val_summary.release_allowed}")

    # ── Step 4: Build config ──────────────────────────────────────────
    config = QuoteConfig(
        quote_ref="QT0000082",
        quote_date=date.today().strftime("%d.%m.%Y"),
        client_name="STEG — Société Tunisienne de l'Electricité et du Gaz",
        client_address="Tunisia",
        project_name="KXT-26011 — Tunisia Substation Project",
        section_label="16.2",
        section_title="LT Cable Termination",
        section_description=(
            "1.1KV End Termination Double compression type Cable glands for 1.1kV grade, "
            "Aluminium conductor, XLPE/PVC insulated, armoured, FRLS PVC sheathed cables "
            "including Lugs, Glands etc."
        ),
        discount_pct=DISCOUNT,
        generated_by="Sales Team",
        approved_by="Director",
        currency="INR",
        is_export=True,
        include_validation_sheet=True,
    )

    # ── Step 5: Generate output ───────────────────────────────────────
    print("  [3/4] Generating Excel output...")
    out_path = "/home/claude/Braco_Quotation_QT0000082.xlsx"
    result = generate_quotation(test_items, results, val_summary, config, out_path)

    if result["ok"]:
        print(f"  [4/4] ✅ File generated: {result['path']}")
    else:
        print(f"  [4/4] ❌ BLOCKED: {result['reason']}")
        return

    # ── Step 6: Test blocked scenario ────────────────────────────────
    print("\n  Testing release gate — injecting a BLOCK...")
    from copy import deepcopy
    from braco_validator import validate as validate_one

    bad_items = deepcopy(test_items[:3])
    bad_results = [run_selection(i) for i in bad_items]
    # Inject a bad price to trigger BLOCK
    if bad_results[0].gland:
        bad_results[0].gland["list_price"] = 9999

    bad_summary = validate_quote(bad_items, bad_results, DISCOUNT)
    blocked_result = generate_quotation(
        bad_items, bad_results, bad_summary, config,
        "/home/claude/should_not_exist.xlsx"
    )
    if not blocked_result["ok"]:
        print(f"  ✅ Release gate HELD. No file created.")
        print(f"     Reason: {blocked_result['reason'][:100]}")
    else:
        print("  ❌ Release gate FAILED — file was generated despite BLOCK")

    # ── Step 7: Recalculate formulas ──────────────────────────────────
    print(f"\n  Recalculating Excel formulas...")
    import subprocess
    r = subprocess.run(
        ["python3", "/home/claude/recalc.py", out_path, "60"],
        capture_output=True, text=True, cwd="/home/claude"
    )
    output = r.stdout.strip()
    if output:
        try:
            import json
            recalc_result = json.loads(output)
            if recalc_result.get("status") == "success":
                print(f"  ✅ Formulas recalculated — {recalc_result.get('total_formulas',0)} formulas, 0 errors")
            else:
                errs = recalc_result.get("error_summary", {})
                print(f"  ⚠️  Formula errors found: {errs}")
        except:
            print(f"  Recalc output: {output[:100]}")
    else:
        print(f"  Recalc: {r.stderr[:100] if r.stderr else 'completed'}")

    print("\n" + "="*68)
    print(f"  OUTPUT: {out_path}")
    print("="*68 + "\n")


if __name__ == "__main__":
    run_output_test()
